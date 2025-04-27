import openpyxl
from io import BytesIO

from django.conf import settings


BOLD_FONT = openpyxl.styles.Font(bold=True)


def to_excel_attachment(title: str, headers: list, rows: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append(headers)

    for row in rows:
        ws.append(row)

    # auto-adjust column widths
    for column_cells in ws.columns:
        max_cell_length = max(
            len(str(cell.value))
            if cell.value else 0 for cell in column_cells
        )
        length = min(max_cell_length, settings.EXCEL_MAX_COLUMN_WIDTH)

        column_letter = openpyxl.utils.get_column_letter(
            column_cells[0].column)
        ws.column_dimensions[column_letter].width = length

    # make first column (ID) small
    ws.column_dimensions['A'].width = 10

    ws.freeze_panes = ws['A2']

    for cell in ws[1]:
        cell.font = BOLD_FONT

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
